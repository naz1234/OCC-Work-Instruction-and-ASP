#!/usr/bin/env python3
"""Extract register values and original embedded hyperlinks from an OCC workbook."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import openpyxl


CONDITIONS = {"normal", "degraded", "emergency"}
REGISTER_SHEET = "OCC Work Instructions"


def cell_text(value: object) -> str:
    """Preserve Excel cell content without inventing labels for blank cells."""
    return "" if value is None else str(value)


def extract_workbook(source_path: Path) -> dict[str, object]:
    workbook = openpyxl.load_workbook(source_path, data_only=False, read_only=False)

    if REGISTER_SHEET not in workbook.sheetnames:
        raise ValueError(f"The workbook must contain a worksheet named {REGISTER_SHEET!r}.")

    sheet = workbook[REGISTER_SHEET]
    current_group = ""
    current_condition = ""
    documents: list[dict[str, object]] = []

    for row_number in range(4, sheet.max_row + 1):
        cells = [sheet.cell(row_number, column) for column in range(1, 6)]
        first, title, reference, line, folder = cells

        if first.value is not None and all(cell.value is None for cell in cells[1:]):
            heading = cell_text(first.value)
            if heading.strip().casefold() in CONDITIONS:
                current_condition = heading
            else:
                current_group = heading
                current_condition = ""
            continue

        # Repeated printed page headers and other workbook metadata are not documents.
        if title.value is None or reference.value is None:
            continue

        if not current_group or not current_condition:
            raise ValueError(
                f"Document on Excel row {row_number} has no preceding group and condition heading."
            )

        documents.append(
            {
                "row": row_number,
                "serial": cell_text(first.value),
                "title": cell_text(title.value),
                "reference": cell_text(reference.value),
                "line": cell_text(line.value),
                "folder": cell_text(folder.value),
                "url": reference.hyperlink.target if reference.hyperlink else "",
                "group": current_group,
                "condition": current_condition,
            }
        )

    if not documents:
        raise ValueError("No document rows were found in the OCC register worksheet.")

    return {
        "metadata": {
            "source": source_path.name,
            "sheet": REGISTER_SHEET,
            "document_number": cell_text(sheet["D1"].value),
            "revision": cell_text(sheet["D2"].value),
            "count": len(documents),
            "hyperlink_count": sum(bool(document["url"]) for document in documents),
        },
        "documents": documents,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to the source Excel workbook.")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parents[1] / "data.js",
        help="Output JavaScript data file; defaults to the website's data.js.",
    )
    arguments = parser.parse_args()

    if not arguments.workbook.is_file():
        parser.error(f"Workbook does not exist: {arguments.workbook}")

    payload = extract_workbook(arguments.workbook)
    serialized = json.dumps(payload, ensure_ascii=False, indent=2)
    arguments.output.write_text(
        "// Generated from the OCC work-instruction workbook.\n"
        "// Run scripts/update_documents.py to refresh this file from a revised workbook.\n"
        f"window.OCC_DATA = {serialized};\n",
        encoding="utf-8",
    )

    metadata = payload["metadata"]
    print(
        f"Generated {arguments.output}: {metadata['count']} documents, "
        f"{metadata['hyperlink_count']} original embedded hyperlinks."
    )


if __name__ == "__main__":
    main()
